from __future__ import annotations
from typing import List, Tuple


class XlsxExtractor:
    """Extract tabular data from .xlsx/.xls files as Markdown tables."""

    def extract(self, path: str) -> List[Tuple[int, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("Install openpyxl: pip install 'docintel[office]'")

        wb = load_workbook(path, read_only=True, data_only=True)
        pages: list[tuple[int, str]] = []

        for sheet_num, sheet in enumerate(wb.worksheets, 1):
            rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if any(cells):
                    rows.append(cells)

            if not rows:
                continue

            # Normalise column count
            max_cols = max(len(r) for r in rows)
            for r in rows:
                r += [""] * (max_cols - len(r))

            lines: list[str] = [f"## Sheet: {sheet.title}\n"]
            lines.append("| " + " | ".join(rows[0]) + " |")
            lines.append("| " + " | ".join(["---"] * max_cols) + " |")
            for row in rows[1:]:
                lines.append("| " + " | ".join(row) + " |")

            # Split into pages of 200 rows so chunks stay manageable
            chunk_size = 200
            header_lines = lines[:2]
            data_lines = lines[2:]
            for i in range(0, max(1, len(data_lines)), chunk_size):
                chunk_lines = header_lines + data_lines[i : i + chunk_size]
                pages.append((sheet_num * 1000 + i // chunk_size, "\n".join(chunk_lines)))

        wb.close()
        return pages if pages else [(1, "")]
